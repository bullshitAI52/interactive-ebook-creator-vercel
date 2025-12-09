# -*- coding: utf-8 -*-
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk, ImageCms
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
from logging.handlers import RotatingFileHandler
import subprocess

# --- 全局函数 ---

# 配置轮转日志，防止日志无限增长
logger = logging.getLogger("image_tool")
logger.setLevel(logging.INFO)
_handler = RotatingFileHandler('image_tool.log', maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
if not logger.handlers:
    logger.addHandler(_handler)

def get_image_info_and_thumbnail(file_path):
    """获取单张图片的详细信息和缩略图"""
    info_dict = {'file_path': file_path, 'error': None, 'thumbnail_bytes': None}
    try:
        with Image.open(file_path) as img:
            width, height = img.size
            dpi = img.info.get('dpi', (72, 72))
            dpi_x = dpi[0] if isinstance(dpi, tuple) and len(dpi) > 0 and dpi[0] > 0 else 72
            dpi_y = dpi[1] if isinstance(dpi, tuple) and len(dpi) > 1 and dpi[1] > 0 else 72
            width_cm = (width / dpi_x) * 2.54
            height_cm = (height / dpi_y) * 2.54
            info_dict.update({
                'pixel_size': (width, height),
                'physical_size': (round(width_cm, 2), round(height_cm, 2)),
                'dpi': (int(dpi_x), int(dpi_y)),
                'color_mode': img.mode,
                'format': img.format,
                'file_size': os.path.getsize(file_path),
            })
            try:
                # 创建缩略图
                thumb_copy = img.copy()
                thumb_copy.thumbnail((128, 128))
                img_byte_arr = BytesIO()
                # 统一保存为PNG以支持透明度且避免JPEG问题
                thumb_copy.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                info_dict['thumbnail_bytes'] = img_byte_arr
            except Exception:
                pass  # 缩略图失败不影响主流程
    except (IOError, OSError, Image.DecompressionBombError) as e_main:
        info_dict['error'] = str(e_main)
        logger.warning(f"图片信息读取失败: {file_path} - {e_main}")
    return os.path.basename(file_path), info_dict

def format_file_size(size_bytes):
    """格式化文件大小显示"""
    if size_bytes is None: return "N/A"
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    return f"{size_bytes / 1024:.2f} KB"

# --- 主程序类 ---

class ImageToolApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("图片信息批量处理工具 v2.2 (最终修正版)  作者@zwm")
        self.root.geometry("1200x800")

        self.folder_path = None
        self.cached_image_info = {}
        self.preview_photo = None
        # 列排序状态
        self._sort_state = {}

        self._setup_ui()

    def _setup_ui(self):
        """搭建用户界面 (已修正Linter报错)"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- 顶部：按钮区 (位置已调整) ---
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 10))
        
        self.select_button = ttk.Button(button_frame, text="选择文件夹", command=self.select_folder_and_load_info)
        self.select_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.rename_button = ttk.Button(button_frame, text="批量重命名", command=self.start_rename_task, state=tk.DISABLED)
        self.rename_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.export_button = ttk.Button(button_frame, text="导出到Excel", command=self.start_export_task, state=tk.DISABLED)
        self.export_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.convert_cmyk_button = ttk.Button(button_frame, text="转为CMYK", command=lambda: self.start_color_convert_task('CMYK'), state=tk.DISABLED)
        self.convert_cmyk_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.convert_rgb_button = ttk.Button(button_frame, text="转为RGB", command=lambda: self.start_color_convert_task('RGB'), state=tk.DISABLED)
        self.convert_rgb_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.refresh_button = ttk.Button(button_frame, text="刷新", command=self.refresh_folder, state=tk.DISABLED)
        self.refresh_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        # --- 中部内容区 ---
        paned_window = ttk.PanedWindow(main_frame, orient=tk.VERTICAL)
        paned_window.pack(fill=tk.BOTH, expand=True)

        # --- 图片列表区 ---
        list_labelframe = ttk.Labelframe(paned_window, text="图片列表", padding="5")
        paned_window.add(list_labelframe, weight=3)

        columns = ('filename', 'pixels', 'size_cm', 'dpi', 'mode', 'filesize')
        self.tree = ttk.Treeview(list_labelframe, columns=columns, show='headings', height=15)
        headings_text = {'filename': '文件名', 'pixels': '像素尺寸', 'size_cm': '物理尺寸(cm)', 
                         'dpi': 'DPI', 'mode': '色彩模式', 'filesize': '文件大小'}
        for col, text in headings_text.items():
            # 列头点击排序
            self.tree.heading(col, text=text, command=lambda c=col: self._on_heading_click(c))
        
        self.tree.column('filename', width=300, stretch=tk.YES)
        for col in columns[1:]:
            self.tree.column(col, width=120, anchor='center')

        list_scrollbar_y = ttk.Scrollbar(list_labelframe, orient=tk.VERTICAL, command=self.tree.yview)
        list_scrollbar_x = ttk.Scrollbar(list_labelframe, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.config(yscrollcommand=list_scrollbar_y.set, xscrollcommand=list_scrollbar_x.set)
        list_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        list_scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)

        # --- 图片预览区 ---
        preview_labelframe = ttk.Labelframe(paned_window, text="图片预览", padding="5")
        paned_window.add(preview_labelframe, weight=2)
        self.preview_label = ttk.Label(preview_labelframe, text="\n\n请在列表中选择图片以预览\n\n", anchor=tk.CENTER, justify=tk.CENTER)
        self.preview_label.pack(fill=tk.BOTH, expand=True)

        # --- 底部：状态栏 ---
        status_frame = ttk.Frame(main_frame, padding="5")
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.status_label = ttk.Label(status_frame, text="欢迎使用！请先选择一个包含图片的文件夹。")
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.progress_bar = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL, length=200, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, padx=5)

    # --- 按钮状态与界面更新 ---

    def set_buttons_state(self, state):
        """统一设置功能按钮的状态"""
        self.rename_button.config(state=state)
        self.export_button.config(state=state)
        self.convert_cmyk_button.config(state=state)
        self.convert_rgb_button.config(state=state)
        self.refresh_button.config(state=state if self.folder_path else tk.DISABLED)
        # "选择文件夹"按钮的状态由其他逻辑独立控制，这里不作处理

    def on_tree_select(self, event):
        """当在列表中选择一项时，更新预览图"""
        selection = self.tree.selection()
        if not selection:
            return
        filename = self.tree.item(selection[0], 'values')[0]
        info = self.cached_image_info.get(filename)
        if info and info.get('thumbnail_bytes'):
            try:
                thumbnail_data = info['thumbnail_bytes']
                thumbnail_data.seek(0)
                pil_image = Image.open(thumbnail_data)
                self.preview_photo = ImageTk.PhotoImage(pil_image)
                self.preview_label.config(image=self.preview_photo, text="")
            except Exception as e:
                self.preview_photo = None
                self.preview_label.config(image='', text=f"无法显示预览:\n{e}")
        else:
            self.preview_photo = None
            self.preview_label.config(image='', text="\n\n无预览可用\n\n")

    def _on_heading_click(self, column):
        # 简单可逆排序
        items = list(self.tree.get_children(''))
        if not items:
            return
        reverse = self._sort_state.get(column, False)
        def keyfunc(item):
            val = self.tree.set(item, column)
            try:
                # 尝试对 "NxM" 或 "AxBcm" 做数值排序（宽+高）
                if column in ('pixels', 'size_cm', 'dpi') and 'x' in val:
                    v = val.replace('cm', '')
                    parts = [p for p in v.split('x') if p]
                    nums = [float(p) for p in parts]
                    return sum(nums)
                if column == 'filesize':
                    # 按 KB 近似比较
                    if val.endswith('MB'):
                        return float(val.split()[0]) * 1024
                    if val.endswith('KB'):
                        return float(val.split()[0])
                # 其他尝试转成数值
                return float(val)
            except Exception:
                return val
        items.sort(key=keyfunc, reverse=reverse)
        self._sort_state[column] = not reverse
        for idx, it in enumerate(items):
            self.tree.move(it, '', idx)

    def update_status(self, message, is_error=False):
        """更新状态栏消息"""
        self.status_label.config(text=message, foreground="red" if is_error else "black")
        self.root.update_idletasks()

    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar['value'] = value
        self.root.update_idletasks()

    def display_image_info_from_cache(self):
        """将缓存的图片信息显示在列表中"""
        self.tree.delete(*self.tree.get_children())
        if not self.folder_path:
            self.update_status("请选择包含图片的文件夹。")
            return
        if not self.cached_image_info:
            self.update_status(f"文件夹 '{os.path.basename(self.folder_path)}' 中没有支持的图片。", is_error=True)
            return
        
        for file_name, info in sorted(self.cached_image_info.items()):
            if info and not info.get('error'):
                values = (file_name,f"{info['pixel_size'][0]}x{info['pixel_size'][1]}",f"{info['physical_size'][0]:.2f}x{info['physical_size'][1]:.2f}",f"{info['dpi'][0]}x{info['dpi'][1]}",info['color_mode'],format_file_size(info['file_size']))
            else:
                values = (file_name, '无法读取', '', '', '', '')
            self.tree.insert('', tk.END, values=values)

    # --- 核心功能：后台任务启动 ---

    def select_folder_and_load_info(self):
        """选择文件夹并加载图片信息"""
        new_folder_path = filedialog.askdirectory(title="请选择图片文件夹")
        if not new_folder_path:
            self.update_status("已取消选择。")
            return
        self.folder_path = new_folder_path
        self.update_status(f"正在加载图片信息: {self.folder_path}")
        self._start_background_load()

    def refresh_folder(self):
        """刷新当前文件夹"""
        if not self.folder_path:
            self.update_status("没有可刷新的文件夹。", is_error=True)
            return
        self.update_status(f"正在刷新: {self.folder_path}")
        self._start_background_load()

    def _start_background_load(self):
        """启动后台线程加载信息（供选择和刷新调用）"""
        def _prep():
            self.tree.delete(*self.tree.get_children())
            self.cached_image_info = {}
            self.preview_label.config(image='', text="\n\n请在列表中选择图片以预览\n\n")
            self.preview_photo = None
            self.update_progress(0)
            self.set_buttons_state(tk.DISABLED)
            self.select_button.config(state=tk.DISABLED)
        self.root.after(0, _prep)
        threading.Thread(target=self._load_info_task, daemon=True).start()

    def _load_info_task(self):
        """[后台线程] 加载图片信息"""
        if not self.folder_path:
            self.root.after(0, self._task_error, "文件夹路径无效。")
            return

        try:
            supported_formats = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif')
            image_files = [
                f for f in os.listdir(self.folder_path)
                if os.path.isfile(os.path.join(self.folder_path, f))
                and not f.startswith('.')
                and f.lower().endswith(supported_formats)
            ]
            
            if not image_files:
                self.root.after(0, self.update_status, f"文件夹 '{os.path.basename(self.folder_path)}' 中没有支持的图片。", True)
                self.root.after(0, self.update_progress, 100)
                return

            temp_info_cache = {}
            max_workers = min(32, (os.cpu_count() or 4) * 5)
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_filename = {
                    executor.submit(get_image_info_and_thumbnail, os.path.join(self.folder_path, fname)): fname 
                    for fname in image_files
                }
                for i, future in enumerate(as_completed(future_to_filename), 1):
                    try:
                        filename, info_data = future.result()
                        temp_info_cache[filename] = info_data
                    except (IOError, OSError, Image.DecompressionBombError) as exc:
                        filename = future_to_filename[future]
                        temp_info_cache[filename] = {'error': str(exc)}
                        logger.warning(f"多线程处理失败: {filename} - {exc}")
                    self.root.after(0, self.update_progress, (i / len(image_files)) * 100)
            
            self.cached_image_info = temp_info_cache
            self.root.after(0, self.display_image_info_from_cache)
            self.root.after(0, self.update_status, f"成功加载 {len(self.cached_image_info)} 张图片。")
        except Exception as e:
            self.root.after(0, self.update_status, f"加载图片信息出错: {e}", True)
            logger.error(f"加载图片信息失败: {e}")
        finally:
            self.root.after(0, lambda: self.select_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.set_buttons_state(tk.NORMAL if self.cached_image_info else tk.DISABLED))

    def start_rename_task(self):
        """启动批量重命名任务"""
        if not self.folder_path or not self.cached_image_info:
            messagebox.showwarning("警告", "请先选择文件夹并加载图片。")
            return
        msg = "此操作将根据图片的物理尺寸重命名文件 (例如, '原文件名_宽x高cm.jpg')。\n\n此操作无法撤销，是否继续？"
        if messagebox.askyesno("确认重命名", msg, icon='warning'):
            self._start_task(self._rename_task, "正在批量重命名...")

    def _rename_task(self):
        """[后台线程] 执行重命名 (已修正路径问题)"""
        if not self.folder_path:
            return self._task_error("未选择文件夹，无法重命名。")

        rename_log, to_rename = [], {}
        current_files = {f for f in os.listdir(self.folder_path) if os.path.isfile(os.path.join(self.folder_path, f))}
        for fname, finfo in self.cached_image_info.items():
            if fname in current_files and finfo and not finfo.get('error') and finfo.get('physical_size'):
                to_rename[fname] = finfo
        
        if not to_rename:
            return self._task_complete("未找到可重命名的有效图片。", [], "重命名", refresh=True)

        for i, (original_name, info) in enumerate(to_rename.items(), 1):
            w_cm, h_cm = info['physical_size']
            base, ext = os.path.splitext(original_name)
            new_name = f"{base}_{round(w_cm)}x{round(h_cm)}cm{ext.lower()}"
            counter = 1
            final_name = new_name
            while final_name in (current_files - {original_name}):
                final_name = f"{base}_{round(w_cm)}x{round(h_cm)}cm_{counter}{ext.lower()}"
                counter += 1
            
            if final_name != original_name:
                try:
                    src_path = os.path.join(self.folder_path, original_name)
                    dst_path = os.path.join(self.folder_path, final_name)
                    os.rename(src_path, dst_path)
                    rename_log.append(f"成功: '{original_name}' -> '{final_name}'")
                    current_files.remove(original_name)
                    current_files.add(final_name)
                except Exception as e:
                    rename_log.append(f"失败: {original_name} - {e}")
            self.root.after(0, self.update_progress, (i / len(to_rename)) * 100)
        self._task_complete("批量重命名完成。", rename_log, "重命名", refresh=True)

    def start_color_convert_task(self, target_mode):
        """启动色彩模式转换任务"""
        if not self.folder_path or not self.cached_image_info:
            messagebox.showwarning("警告", "请先选择文件夹并加载图片。")
            return
        output_path = os.path.join(self.folder_path, f"{target_mode}_Converted")
        msg = f"图片将转换为 {target_mode} 并保存在新子文件夹:\n'{output_path}'\n\n原始文件不会被修改，是否继续？"
        if messagebox.askyesno("确认转换", msg):
            self._start_task(self._convert_task, f"正在转换为 {target_mode}...", target_mode, output_path)

    def _convert_task(self, target_mode, output_path):
        """[后台线程] 执行颜色转换 (已优化)"""
        if not self.folder_path:
            return self._task_error("未选择文件夹，无法转换。")
            
        os.makedirs(output_path, exist_ok=True)
        to_convert = {f: i for f, i in self.cached_image_info.items() if i and not i.get('error')}
        if not to_convert:
            return self._task_complete("没有找到可以转换的有效图片。", [], f"转换为 {target_mode}")

        log, count = [], 0
        for i, (fname, info) in enumerate(to_convert.items(), 1):
            if info.get('color_mode') == target_mode:
                log.append(f"跳过: '{fname}' 已是 {target_mode}。")
                continue
            
            original_path = os.path.join(self.folder_path, fname)
            base, orig_ext = os.path.splitext(fname)
            output_ext = '.jpg' if target_mode == 'CMYK' else orig_ext
            output_file = os.path.join(output_path, base + output_ext)
            
            try:
                with Image.open(original_path) as img:
                    # 处理带透明度的图片
                    if img.mode in ('RGBA', 'LA') and target_mode != 'RGBA':
                        background = Image.new('RGB', img.size, (255, 255, 255))
                        background.paste(img, mask=img.split()[-1])
                        img = background

                    # 改进的RGB转CMYK算法，更接近Photoshop的效果
                    if target_mode == 'CMYK':
                        converted_img = self._rgb_to_cmyk_improved(img)
                    else:
                        converted_img = img.convert(target_mode)
                    
                    if output_ext.lower() in ['.jpg', '.jpeg']:
                        converted_img.save(output_file, 'jpeg', quality=95)
                    else:
                        converted_img.save(output_file)
                        
                    log.append(f"成功: '{fname}' -> '{os.path.basename(output_file)}'")
                    count += 1
            except Exception as e:
                log.append(f"失败: '{fname}' - {e}")
            self.root.after(0, self.update_progress, (i / len(to_convert)) * 100)
        self._task_complete(f"转换完成，{count} 张图片已处理。", log, f"转换为 {target_mode}")

    def _rgb_to_cmyk_improved(self, rgb_img):
        """改进的RGB转CMYK算法，更接近Photoshop的效果"""
        # 方法1: 尝试使用ImageMagick (效果最好)
        try:
            return self._rgb_to_cmyk_imagemagick(rgb_img)
        except Exception as e:
            logger.debug(f"ImageMagick转换失败: {e}")
            
        # 方法2: 尝试使用ICC色彩配置文件
        try:
            return self._rgb_to_cmyk_with_icc(rgb_img)
        except Exception as e:
            logger.debug(f"ICC转换失败: {e}")
            
        # 方法3: 使用改进的数学算法
        return self._rgb_to_cmyk_math(rgb_img)
    
    def _rgb_to_cmyk_with_icc(self, rgb_img):
        """使用ICC色彩配置文件进行转换"""
        if rgb_img.mode != 'RGB':
            rgb_img = rgb_img.convert('RGB')
        
        # 尝试使用系统ICC配置文件
        try:
            # 使用sRGB到CMYK的转换
            # 注意：这需要系统有相应的ICC配置文件
            cmyk_img = ImageCms.profileToProfile(
                rgb_img, 
                'sRGB',  # 输入配置文件
                'USWebCoatedSWOP',  # 输出配置文件（印刷常用）
                outputMode='CMYK'
            )
            return cmyk_img
        except Exception as e:
            # 如果ICC转换失败，回退到数学算法
            logger.debug(f"ICC转换失败，使用数学算法: {e}")
            return self._rgb_to_cmyk_math(rgb_img)
    
    def _rgb_to_cmyk_math(self, rgb_img):
        """改进的数学转换算法"""
        if rgb_img.mode != 'RGB':
            rgb_img = rgb_img.convert('RGB')
        
        # 使用numpy优化性能（如果可用）
        try:
            import numpy as np
            return self._rgb_to_cmyk_numpy(rgb_img)
        except ImportError:
            # 回退到逐像素处理
            rgb_data = rgb_img.getdata()
            cmyk_data = []
            
            for r, g, b in rgb_data:
                # 转换为0-1范围的浮点数
                r_norm = r / 255.0
                g_norm = g / 255.0
                b_norm = b / 255.0
                
                # 计算CMY值
                c = 1.0 - r_norm
                m = 1.0 - g_norm
                y = 1.0 - b_norm
                
                # 计算黑版 (K) - 使用最大值法
                k = min(c, m, y)
                
                # 调整CMY值，考虑黑版
                if k < 1.0:
                    c = (c - k) / (1.0 - k)
                    m = (m - k) / (1.0 - k)
                    y = (y - k) / (1.0 - k)
                else:
                    c = m = y = 0.0
                
                # 转换为0-255范围的整数
                c_int = int(c * 255)
                m_int = int(m * 255)
                y_int = int(y * 255)
                k_int = int(k * 255)
                
                cmyk_data.append((c_int, m_int, y_int, k_int))
            
            cmyk_img = Image.new('CMYK', rgb_img.size)
            cmyk_img.putdata(cmyk_data)
            return cmyk_img
    
    def _rgb_to_cmyk_numpy(self, rgb_img):
        """使用numpy优化的RGB转CMYK算法"""
        import numpy as np
        
        rgb_array = np.array(rgb_img, dtype=np.float32) / 255.0
        
        # 计算CMY
        cmy = 1.0 - rgb_array
        
        # 计算黑版 (K)
        k = np.min(cmy, axis=2, keepdims=True)
        
        # 调整CMY值
        mask = k < 1.0
        cmy_adjusted = np.where(mask, (cmy - k) / (1.0 - k), 0.0)
        
        # 合并CMYK
        cmyk_array = np.concatenate([cmy_adjusted, k], axis=2)
        
        # 转换为0-255整数
        cmyk_array = (cmyk_array * 255).astype(np.uint8)
        
        return Image.fromarray(cmyk_array, 'CMYK')
    
    def _rgb_to_cmyk_photoshop_like(self, rgb_img):
        """更接近Photoshop的转换算法"""
        if rgb_img.mode != 'RGB':
            rgb_img = rgb_img.convert('RGB')
        
        # 使用numpy优化性能（如果可用）
        try:
            import numpy as np
            return self._rgb_to_cmyk_photoshop_numpy(rgb_img)
        except ImportError:
            # 回退到逐像素处理
            rgb_data = rgb_img.getdata()
            cmyk_data = []
            
            for r, g, b in rgb_data:
                # 转换为0-1范围的浮点数
                r_norm = r / 255.0
                g_norm = g / 255.0
                b_norm = b / 255.0
                
                # 计算CMY值
                c = 1.0 - r_norm
                m = 1.0 - g_norm
                y = 1.0 - b_norm
                
                # Photoshop风格的黑版计算
                # 使用更复杂的黑版生成算法
                k = min(c, m, y)
                
                # 应用总墨量限制 (类似Photoshop的设置)
                total_ink = c + m + y + k
                max_total_ink = 3.0  # 最大总墨量限制
                
                if total_ink > max_total_ink:
                    scale = max_total_ink / total_ink
                    c *= scale
                    m *= scale
                    y *= scale
                    k *= scale
                
                # 转换为0-255范围的整数
                c_int = int(c * 255)
                m_int = int(m * 255)
                y_int = int(y * 255)
                k_int = int(k * 255)
                
                cmyk_data.append((c_int, m_int, y_int, k_int))
            
            cmyk_img = Image.new('CMYK', rgb_img.size)
            cmyk_img.putdata(cmyk_data)
            return cmyk_img
    
    def _rgb_to_cmyk_photoshop_numpy(self, rgb_img):
        """使用numpy优化的Photoshop风格转换算法"""
        import numpy as np
        
        rgb_array = np.array(rgb_img, dtype=np.float32) / 255.0
        
        # 计算CMY
        cmy = 1.0 - rgb_array
        
        # 计算黑版 (K)
        k = np.min(cmy, axis=2, keepdims=True)
        
        # 调整CMY值
        mask = k < 1.0
        cmy_adjusted = np.where(mask, (cmy - k) / (1.0 - k), 0.0)
        
        # 应用总墨量限制
        total_ink = np.sum(cmy_adjusted, axis=2, keepdims=True) + k
        max_total_ink = 3.0
        mask_overflow = total_ink > max_total_ink
        scale = np.where(mask_overflow, max_total_ink / total_ink, 1.0)
        
        cmyk_array = np.concatenate([cmy_adjusted * scale, k * scale], axis=2)
        
        # 转换为0-255整数
        cmyk_array = (cmyk_array * 255).astype(np.uint8)
        
        return Image.fromarray(cmyk_array, 'CMYK')
    
    def _rgb_to_cmyk_imagemagick(self, rgb_img):
        """使用ImageMagick进行高质量RGB转CMYK转换"""
        import tempfile
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_input:
            with tempfile.NamedTemporaryFile(suffix='.tif', delete=False) as temp_output:
                temp_input_path = temp_input.name
                temp_output_path = temp_output.name
        
        try:
            # 保存RGB图像
            rgb_img.save(temp_input_path)
            
            # 使用ImageMagick转换
            # -colorspace CMYK: 转换为CMYK色彩空间
            # 先尝试不使用ICC配置文件，避免文件不存在错误
            cmd_simple = [
                'magick', temp_input_path,
                '-colorspace', 'CMYK',
                temp_output_path
            ]
            
            # 执行命令
            result = subprocess.run(cmd_simple, capture_output=True, text=True, timeout=30)
            
            if result.returncode != 0:
                raise Exception(f"ImageMagick转换失败: {result.stderr}")
            
            # 读取转换后的图像
            cmyk_img = Image.open(temp_output_path)
            
            return cmyk_img
            
        finally:
            # 清理临时文件
            try:
                os.unlink(temp_input_path)
            except:
                pass
            try:
                os.unlink(temp_output_path)
            except:
                pass

    def start_export_task(self):
        """启动导出到Excel任务"""
        if not self.folder_path or not self.cached_image_info:
            messagebox.showwarning("警告", "请先选择文件夹并加载图片。")
            return
        default_name = f"{os.path.basename(self.folder_path or '图片')}_报告.xlsx"
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")], title="保存Excel报告", initialfile=default_name, initialdir=self.folder_path)
        if save_path:
            self._start_task(self._excel_task, "正在导出到Excel...", save_path)

    def _excel_task(self, save_path):
        """[后台线程] 执行导出到Excel (已修正)"""
        if not self.folder_path:
            return self._task_error("未选择文件夹，无法导出。")

        wb = Workbook()
        ws = wb.active
        if ws is None: # 防御性编程，确保工作表存在
            ws = wb.create_sheet("图片信息报告")
        ws.title = "图片信息报告"
        
        headers = ["文件名", "像素尺寸", "物理尺寸(cm)", "DPI", "色彩模式", "文件大小"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        for i, (fname, info) in enumerate(sorted(self.cached_image_info.items()), 1):
            if info and not info.get('error'):
                row = [
                    fname,
                    f"{info['pixel_size'][0]}x{info['pixel_size'][1]}",
                    f"{info['physical_size'][0]:.2f}x{info['physical_size'][1]:.2f}",
                    f"{info['dpi'][0]}x{info['dpi'][1]}",
                    info['color_mode'],
                    format_file_size(info['file_size'])
                ]
            else:
                row = [fname, '无法读取', '', '', '', '']
            ws.append(row)
            self.root.after(0, self.update_progress, (i / len(self.cached_image_info)) * 100)
        
        wb.save(save_path)
        self._task_complete(f"成功导出到: {save_path}", [], "导出Excel")

    # --- 后台任务管理 ---
    
    def _start_task(self, task_func, status_msg, *args):
        """通用任务启动器"""
        self.update_status(status_msg)
        self.update_progress(0)
        self.set_buttons_state(tk.DISABLED)
        self.select_button.config(state=tk.DISABLED)
        threading.Thread(target=self._task_runner, args=(task_func, args), daemon=True).start()

    def _task_runner(self, task_func, args):
        """通用任务运行器，捕获异常"""
        try:
            task_func(*args)
        except Exception as e:
            self.root.after(0, self._task_error, f"发生意外错误: {e}")

    def _task_complete(self, summary, log, op_name, refresh=False):
        """任务完成后的回调 (已修正f-string兼容性)"""
        self.root.after(0, self.update_status, summary)
        self.root.after(0, lambda: self.set_buttons_state(tk.NORMAL if self.cached_image_info else tk.DISABLED))
        self.root.after(0, lambda: self.select_button.config(state=tk.NORMAL))
        self.root.after(0, self.update_progress, 100)
        
        details = f"详细日志:\n" + "\n".join(log) if log else ""
        full_message = f"{summary}\n\n{details}"
        messagebox.showinfo(f"{op_name}结果", full_message)
        
        if refresh:
            self.root.after(100, self.refresh_folder)

    def _task_error(self, error_msg):
        """任务出错后的回调"""
        self.root.after(0, self.update_status, error_msg, True)
        self.root.after(0, lambda: self.set_buttons_state(tk.NORMAL if self.cached_image_info else tk.DISABLED))
        self.root.after(0, lambda: self.select_button.config(state=tk.NORMAL))
        self.root.after(0, self.update_progress, 0)
        messagebox.showerror("任务出错", error_msg)

# --- 程序入口 ---
if __name__ == "__main__":
    # 强制使用标准 Tkinter 以排除主题库的渲染问题
    root = tk.Tk()
    
    app = ImageToolApp(root)
    root.mainloop()
